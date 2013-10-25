# Parses Plans & Benefits Template and Rate Table files for health
# exchanges. Tested on files from DCHBX, the DC health exchange authority.
# Unzip the .zip file from DCHBX.
#
# python parse_plan_template_files.py ../path/to/files > plans.json

import sys, glob
from datetime import datetime

def main():
	path = sys.argv[1]

	plans = { }

	for fn in glob.glob(path + "/*/Individual/*.xlsm"):
		process_plan_benefits(plans, fn)

	for fn in glob.glob(path + "/*/Individual/*.xls"):
		process_plan_rates(plans, fn)

	import json
	print json.dumps(plans, indent=2, sort_keys=True)

def process_plan_benefits(plans, filename):
	from openpyxl import load_workbook
	wb = load_workbook(filename)
	for sheet_name in wb.get_sheet_names():
		sheet = wb.get_sheet_by_name(sheet_name)
		if sheet_name.startswith("Benefits Package "):
			process_plan_benefits_package(plans, sheet)
		elif sheet_name.startswith("Cost Share Variances "):
			process_plan_costs(plans, sheet)
		elif sheet_name in ("EnableMacros", "DefaultBP", "DefaultCSV", "Names", "Sheet1"):
			pass
		else:
			print "skipping", sheet_name

def process_plan_benefits_package(plans, sheet):
	if sheet.cell("A1").value not in ("Plans & Benefits Template v1.31", "Plans & Benefits Template v1.32"):
		raise Exception("Invalid sheet: " + sheet.cell("A1").value)

	if sheet.cell("A59").value != "Benefit Information":
		raise Exception("Invalid sheet: " + sheet.cell("A59").value)

	hios_issuer = sheet.cell("B2").value
	issuer_state = sheet.cell("B3").value

	market_coverage = sheet.cell("B4").value
	if market_coverage not in ("Individual", ): raise Exception("Invalid market coverage value: " + market_coverage)

	is_dental_only = sheet.cell("B5").value
	if is_dental_only not in ("No", "Yes"): raise Exception("Invalid dental only plan value: " + is_dental_only)
	is_dental_only = (is_dental_only == "Yes")

	tin = sheet.cell("B6").value

	plan_list = []

	cols1 = ("hios_product_id", "hpid", "network_id", "service_area_id", "formulary_id", "new_or_existing_plan", "plan_type",
		"metalic_level", "unique_plan_design", "qhp", "notice_required_for_pregnancy", "referral_required_for_specialist", "specialists_requiring_referral",
		"plan_level_exclusions", "limited_cost_sharing_plan_variation_est_adv_payment", "hsa_elligible", "hsa_hra_employer_contribution",
		"hsa_hra_employer_contribution_amount", "child_only_offering", "child_only_plan_id", "wellness_program_offered", "disease_management_program_offered",
		"ehb_apportionment_for_pediatric_dental", "guaranteed_estimated_rate", "maximum_coinsurance_specialty_drugs", "max_days_inpatient_copay",
		"primary_care_cost_sharing_after_visits", "primary_care_deductible_after_copays", "plan_effective_date", "plan_expiration_date",
		"out_of_country_coverage", "out_of_country_coverage_decription", "out_of_service_area_coverage", "out_of_service_area_coverage_description",
		"national_network", "sbc_url", "enrollment_payment_url", "brochure_url")

	cols2 = ("ehb", "state_mandate", "is_covered", "quantitative_limit", "limit_quantity", "limit_unit", "minimum_stay", "exclusions",
		"explanation", "ehb_variance_reason", "subject_to_deductible_t1", "subject_to_deductible_t2", "excluded_from_in_network_moop",
		"excluded_from_out_of_network_moop")

	# Top metadata by plan

	for r in xrange(8, 58):
		hios_plan = sheet.cell(row=r, column=0).value
		if hios_plan is None: continue

		plan_name = sheet.cell(row=r, column=1).value
		
		plan = {
			"id": hios_plan,
			"name": plan_name,
			"market": market_coverage,
			"is_dental_only": is_dental_only,
			"issuer": {
				"id": hios_issuer,
				"issuer_state": issuer_state,
				"tin": tin,
			},
			"metadata": { },
		}

		for i, col in enumerate(cols1):
			v = sheet.cell(row=r, column=i+2).value
			if v is not None:
				if isinstance(v, datetime): v = v.isoformat()
				plan["metadata"][col] = v

		plan_list.append(plan)

	# Bottom coverage data that seems to apply to all named plans
	coverage = { }
	for r in xrange(60, 162):
		benefit = sheet.cell(row=r, column=0).value
		if benefit is None: continue

		coverage[benefit] = { }

		for i, col in enumerate(cols2):
			v = sheet.cell(row=r, column=i+2).value
			if v is not None:
				coverage[benefit][col] = v

	# Apply the coverage table to all plans and add each plan to the
	# plans dict.
	for plan in plan_list:
		plan["coverage"] = coverage
		plans[plan["id"]] = plan

def process_plan_costs(plans, sheet):
	nrows = sheet.get_highest_row()
	ncols = sheet.get_highest_column()

	for r in xrange(3, nrows):
		hios_plan, variant_id = sheet.cell(row=r, column=0).value.split("-")
		metalic_level = sheet.cell(row=r, column=2).value
		variant_name = sheet.cell(row=r, column=3).value

		if hios_plan not in plans: raise Exception("Invalid plan found in cost table.")
		
		stack = [{}]
		def push(i, k):
			while len(stack) > i: stack.pop(-1)
			stack.append(stack[-1].setdefault(k, {}))

		for c in xrange(4, ncols+1):
			k1 = sheet.cell(row=0, column=c).value
			k2 = sheet.cell(row=1, column=c).value
			k3 = sheet.cell(row=2, column=c).value
			if k1: push(1, k1)
			if k2: push(2, k2)
			v = sheet.cell(row=r, column=c).value
			if v not in (None, ""):
				stack[-1][k3] = v

		costs = stack[0]
		plan = plans[hios_plan]

		if r == 3:
			# The first variant is the standard plan without any cost sharing with
			# the government. For simplicity, we'll embed these costs at the top
			# level of the plan.
			if variant_id != "01": raise Exception("First variant should be 01?")
			if variant_name != ("Standard %s On Exchange Plan" % metalic_level): raise Exception("First variant invalid name: " + variant_name)

			plan["deductibles"] = { }
			plan["maximums"] = { }
			plan["cost_structure"] = costs

			for k in list(costs): # clone
				if k in plan["coverage"]:
					# Some cost information corresponds to covered benefits. Move
					# that information into the coverage section.
					plan["coverage"][k]["costs"] = costs[k]
					del costs[k]
				elif "Deductible" in k:
					# Move deductible-related information into its own category.
					plan["deductibles"][k] = costs[k]
					del costs[k]
				elif "Maximum" in k:
					# Move maximum out of pocket information into its own category.
					plan["maximums"][k] = costs[k]
					del costs[k]

		else:
			# Other variants are for cost sharing. Record the information in a different way.
			plan.setdefault("variants", {})[variant_id] = {
				"id": variant_id,
				"name": variant_name,
				"costs": costs
			}


def process_plan_rates(plans, filename):
	from xlrd import open_workbook, xldate_as_tuple

	wb = open_workbook(filename)
	ws = wb.sheet_by_name("Rate Table")
	if ws.cell(0,0).value not in ("Rates Table Template v2.2", "Rates Table Template v2.3"): raise Exception("Invalid rates table: " + ws.cell(0,0).value + " in " + filename)

	rate_effective_date = datetime(*xldate_as_tuple(ws.cell(7, 1).value, wb.datemode)).isoformat()
	rate_expiration_date = datetime(*xldate_as_tuple(ws.cell(8, 1).value, wb.datemode)).isoformat()

	for r in xrange(13, ws.nrows):
		plan = ws.cell(r, 0).value
		plan_rates = plans[plan].setdefault("rates", [])
		plan_rates.append({
			"rating_area": ws.cell(r, 1).value,
			"tobacco_use": ws.cell(r, 2).value,
			"age": ws.cell(r, 3).value,
			"rate": ws.cell(r, 4).value,
			"effective": rate_effective_date, # applies to all rates, so this isn't the right place to encode it
			"expires": rate_expiration_date, # applies to all rates, so this isn't the right place to encode it
		})

main()
